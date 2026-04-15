# app.py
import os, json
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, flash, Response
from werkzeug.security import generate_password_hash, check_password_hash
from markupsafe import Markup
import smtplib
from email.mime.text import MIMEText




app = Flask(__name__)
app.secret_key = "very_secret_key_here"

##DB_PATH = os.environ.get("DATABASE_PATH", "database.db")


# ---------------- DB ---------------- #
def get_db():
    db_url = os.environ["DATABASE_URL"]

    sslmode = "require" if "render.com" in db_url else "disable"

    return psycopg2.connect(
        db_url,
        cursor_factory=RealDictCursor,
        sslmode=sslmode
    )

import openpyxl
from io import BytesIO

@app.route("/export/patients")
def export_patients():
    if session.get("role") != "doctor":
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Users ──
    cur.execute("SELECT id, username, role, full_name FROM users")
    rows = cur.fetchall()
    ws = wb.create_sheet("Users")
    if rows:
        ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append(list(r.values()))

    # ── Sheet 2: Patient Profiles ──
    cur.execute("SELECT * FROM patient_profiles")
    rows = cur.fetchall()
    ws = wb.create_sheet("Patient Profiles")
    if rows:
        ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append([str(v) if not isinstance(v, (int, float, type(None))) else v for v in r.values()])

    # ── Sheet 3: Patient History ──
    cur.execute("SELECT * FROM patient_history")
    rows = cur.fetchall()
    ws = wb.create_sheet("Patient History")
    if rows:
        ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append([str(v) if not isinstance(v, (int, float, type(None))) else v for v in r.values()])

    # ── Sheet 4: Symptoms (flattened JSON) ──
    # ── Sheet 4: Symptoms (flattened JSON) ──
    cur.execute("""
        SELECT
            u.full_name        AS "Patient Name",
            p.email            AS "Email",
            p.phone            AS "Phone",
            p.gender           AS "Gender",
            p.dob              AS "Date of Birth",
            p.hospital_number  AS "Hospital Number",
            s.created_at       AS "Record Date",
            s.tnss             AS "TNSS Score",
            s.avg_vas          AS "VAS Average",
            s.pattern          AS "Pattern",
            s.follow_up        AS "Follow Up Level",
            s.medicine_effect  AS "Medicine Effect",
            s.recommendation   AS "Recommendation",
            s.raw_form         AS "raw_form",
            s.doctor_notes_updated_at AS "Doctor Notes Updated",
            s.chlorpheniramine, s.other_1st_gen,
            s.cetirizine, s.levocetirizine, s.fexofenadine,
            s.loratadine, s.desloratadine, s.bilastine, s.rupatadine, s.other_2nd_gen,
            s.pseudoephedrine, s.other_oral_decongestant,
            s.triprolidine_pseudo, s.chlorphen_pseudo, s.loratadine_pseudo,
            s.montelukast, s.immunotherapy_oral,
            s.beclomethasone, s.budesonide, s.fluticasone_propionate,
            s.fluticasone_prop_azelastine, s.fluticasone_furoate,
            s.mometasone, s.triamcinolone, s.other_incs,
            s.ephedrine, s.oxymetazoline, s.other_intranasal_decongestant,
            s.azelastine, s.levocabastin, s.ketotifen, s.prednisolone,
            s.nasal_irrigation, s.other_medications,
            s.immunotherapy_inject, s.anti_ige, s.dupilumab, s.benralizumab,
            s.patient_advice, s.next_visit
        FROM users u
        LEFT JOIN patient_profiles p ON u.id = p.user_id
        LEFT JOIN symptoms s ON u.id = s.user_id
        WHERE u.role = 'patient' AND s.id IS NOT NULL
        ORDER BY u.full_name, s.created_at
    """)
    rows = cur.fetchall()
    conn.close()

    ws = wb.create_sheet("Symptoms")

    json_keys = [
        "vas_score1", "vas_score2", "vas_score3",
        "symptom_frequency",
        "Frequently sneeze", "Stuffed nose", "runny nose", "itchy nose",
        "itchy_eyes", "watery_eyes", "itchy_throat", "sore_throat",
        "fatigue", "poor_sleep", "daytime_sleepiness", "snoring",
        "headache", "dry_mouth", "mouth_breathing", "chronic_cough",
        "phlegm_throat", "loss_of_smell", "other_symptom",
        "antihistamine_type", "incs_type", "other_medicine_name"
    ]

    # Friendly display names for JSON keys
    json_labels = [
        "VAS Score 1", "VAS Score 2", "VAS Score 3",
        "Symptom Frequency (days/week)",
        "Sneeze Score", "Stuffed Nose Score", "Runny Nose Score", "Itchy Nose Score",
        "Itchy Eyes", "Watery Eyes", "Itchy Throat", "Sore Throat",
        "Fatigue", "Poor Sleep", "Daytime Sleepiness", "Snoring",
        "Headache", "Dry Mouth", "Mouth Breathing", "Chronic Cough",
        "Phlegm in Throat", "Loss of Smell", "Other Symptom",
        "Antihistamine Used", "Nasal Steroid Used", "Other Medicine"
    ]

    doctor_headers = [
        "Doctor Notes Updated",
        "Chlorpheniramine", "Other 1st Gen",
        "Cetirizine", "Levocetirizine", "Fexofenadine",
        "Loratadine", "Desloratadine", "Bilastine", "Rupatadine", "Other 2nd Gen",
        "Pseudoephedrine", "Other Oral Decongestant",
        "Triprolidine+Pseudoephedrine", "Chlorpheniramine+Pseudoephedrine", "Loratadine+Pseudoephedrine",
        "Montelukast", "Immunotherapy (Oral)",
        "Beclomethasone", "Budesonide", "Fluticasone Propionate",
        "Fluticasone Prop/Azelastine", "Fluticasone Furoate",
        "Mometasone", "Triamcinolone", "Other INCS",
        "Ephedrine", "Oxymetazoline", "Other Intranasal Decongestant",
        "Azelastine", "Levocabastin", "Ketotifen", "Prednisolone",
        "Nasal Irrigation", "Other Medications",
        "Immunotherapy (Inject)", "Anti-IgE", "Dupilumab", "Benralizumab",
        "Patient Advice", "Next Visit"
    ]

    base_headers = [
        "Patient Name", "Email", "Phone", "Gender", "Date of Birth",
        "Hospital Number", "Record Date", "TNSS Score", "VAS Average",
        "Pattern", "Follow Up Level", "Medicine Effect", "Recommendation"
    ]

    ws.append(base_headers + json_labels + doctor_headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for r in rows:
        base_values = [
            r["Patient Name"], r["Email"], r["Phone"], r["Gender"],
            str(r["Date of Birth"]) if r["Date of Birth"] else "",
            r["Hospital Number"],
            str(r["Record Date"]) if r["Record Date"] else "",
            r["TNSS Score"], r["VAS Average"], r["Pattern"],
            r["Follow Up Level"], r["Medicine Effect"], r["Recommendation"]
        ]

        raw = r["raw_form"] or {}
        if isinstance(raw, str):
            raw = json.loads(raw)

        json_values = []
        for key in json_keys:
            val = raw.get(key, "")
            if isinstance(val, list):
                val = ", ".join(val)
            json_values.append(val)

        doctor_values = [
            str(r.get("Doctor Notes Updated")) if r.get("Doctor Notes Updated") else "",
            r.get("chlorpheniramine") or "", r.get("other_1st_gen") or "",
            r.get("cetirizine") or "", r.get("levocetirizine") or "", r.get("fexofenadine") or "",
            r.get("loratadine") or "", r.get("desloratadine") or "", r.get("bilastine") or "",
            r.get("rupatadine") or "", r.get("other_2nd_gen") or "",
            r.get("pseudoephedrine") or "", r.get("other_oral_decongestant") or "",
            r.get("triprolidine_pseudo") or "", r.get("chlorphen_pseudo") or "", r.get("loratadine_pseudo") or "",
            r.get("montelukast") or "", r.get("immunotherapy_oral") or "",
            r.get("beclomethasone") or "", r.get("budesonide") or "", r.get("fluticasone_propionate") or "",
            r.get("fluticasone_prop_azelastine") or "", r.get("fluticasone_furoate") or "",
            r.get("mometasone") or "", r.get("triamcinolone") or "", r.get("other_incs") or "",
            r.get("ephedrine") or "", r.get("oxymetazoline") or "", r.get("other_intranasal_decongestant") or "",
            r.get("azelastine") or "", r.get("levocabastin") or "", r.get("ketotifen") or "",
            r.get("prednisolone") or "", r.get("nasal_irrigation") or "", r.get("other_medications") or "",
            r.get("immunotherapy_inject") or "", r.get("anti_ige") or "",
            r.get("dupilumab") or "", r.get("benralizumab") or "",
            r.get("patient_advice") or "", r.get("next_visit") or "",
        ]

        ws.append(base_values + json_values + doctor_values)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=patient_database_export.xlsx"}
    )






def init_db():
    conn = get_db()
    cur = conn.cursor()

    # USERS
    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL,
        full_name TEXT
    )
    """)

    # PATIENT PROFILE
    cur.execute("""
    CREATE TABLE IF NOT EXISTS patient_profiles (
        id SERIAL PRIMARY KEY,
        user_id INTEGER UNIQUE REFERENCES users(id) ON DELETE CASCADE,
        email TEXT,
        phone TEXT,
        address TEXT,
        dob DATE,
        gender TEXT,
        emergency_contact TEXT,
        insurance_provider TEXT,
        hospital_number TEXT
    )
    """)


    # SYMPTOMS
    cur.execute("""
    CREATE TABLE IF NOT EXISTS symptoms (
        id SERIAL PRIMARY KEY,
        user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
        tnss INTEGER,
        avg_vas REAL,
        pattern TEXT,
        recommendation TEXT,
        follow_up INTEGER DEFAULT 0,
        created_at TIMESTAMP,
        raw_form JSONB,
        medicine_effect INTEGER,
        email_sent BOOLEAN DEFAULT FALSE
    )
    """)

    # Add doctor medication columns to symptoms table
    doctor_cols = [
        ("chlorpheniramine", "TEXT"),
        ("other_1st_gen", "TEXT"),
        ("cetirizine", "TEXT"),
        ("levocetirizine", "TEXT"),
        ("fexofenadine", "TEXT"),
        ("loratadine", "TEXT"),
        ("desloratadine", "TEXT"),
        ("bilastine", "TEXT"),
        ("rupatadine", "TEXT"),
        ("other_2nd_gen", "TEXT"),
        ("pseudoephedrine", "TEXT"),
        ("other_oral_decongestant", "TEXT"),
        ("triprolidine_pseudo", "TEXT"),
        ("chlorphen_pseudo", "TEXT"),
        ("loratadine_pseudo", "TEXT"),
        ("montelukast", "TEXT"),
        ("immunotherapy_oral", "TEXT"),
        ("beclomethasone", "TEXT"),
        ("budesonide", "TEXT"),
        ("fluticasone_propionate", "TEXT"),
        ("fluticasone_prop_azelastine", "TEXT"),
        ("fluticasone_furoate", "TEXT"),
        ("mometasone", "TEXT"),
        ("triamcinolone", "TEXT"),
        ("other_incs", "TEXT"),
        ("ephedrine", "TEXT"),
        ("oxymetazoline", "TEXT"),
        ("other_intranasal_decongestant", "TEXT"),
        ("azelastine", "TEXT"),
        ("levocabastin", "TEXT"),
        ("ketotifen", "TEXT"),
        ("prednisolone", "TEXT"),
        ("nasal_irrigation", "TEXT"),
        ("other_medications", "TEXT"),
        ("immunotherapy_inject", "TEXT"),
        ("anti_ige", "TEXT"),
        ("dupilumab", "TEXT"),
        ("benralizumab", "TEXT"),
        ("patient_advice", "TEXT")
        ("olopatadine_mometasone", "TEXT"),
        ("mepolizumab", "TEXT"),
        ("next_visit", "TEXT"),
        ("doctor_notes_updated_at", "TIMESTAMP"),
    ]
    for col_name, col_type in doctor_cols:
        cur.execute(f"ALTER TABLE symptoms ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
    # PATIENT HISTORY
    cur.execute("""
    CREATE TABLE IF NOT EXISTS patient_history (
        id SERIAL PRIMARY KEY,
        user_id INTEGER UNIQUE REFERENCES users(id) ON DELETE CASCADE,

        symptom_year_pattern TEXT,

        season_summer BOOLEAN,
        season_rainy BOOLEAN,
        season_winter BOOLEAN,
        season_summer_rainy BOOLEAN,
        season_rainy_winter BOOLEAN,
        season_uncertain BOOLEAN,

        duration_per_year TEXT,
        weekly_frequency TEXT,

        time_6_12 BOOLEAN,
        time_12_18 BOOLEAN,
        time_18_24 BOOLEAN,
        time_24_6 BOOLEAN,
        time_all_day BOOLEAN,
        time_uncertain BOOLEAN,

        living_area TEXT,
        near_road BOOLEAN,
        housing_type TEXT,
        air_conditioner BOOLEAN,

        pet_cat BOOLEAN,
        pet_dog BOOLEAN,
        pet_bird BOOLEAN,
        pet_other TEXT,

        trigger_dust BOOLEAN,
        trigger_pollen BOOLEAN,
        trigger_animal BOOLEAN,
        trigger_smoke BOOLEAN,
        trigger_cold_air BOOLEAN,
        trigger_pollution BOOLEAN,
        trigger_stress BOOLEAN,
        trigger_other TEXT,

        smoking_status TEXT,
        cigarettes_per_day INTEGER,
        quit_years INTEGER,
        secondhand_smoke TEXT,

        drug_allergy TEXT,
        drug_allergy_name TEXT,
        drug_allergy_symptom TEXT,
        food_allergy TEXT,
        food_allergy_name TEXT,
        food_allergy_symptom TEXT,

        natural_allergy TEXT,
        natural_allergy_symptom TEXT,

        family_asthma TEXT,
        family_rhinitis TEXT,
        family_allergic_conjunctivitis TEXT,
        family_atopic_dermatitis TEXT,

        work_performance TEXT,
        physical_activity_problem TEXT,
        stairs_problem TEXT,

        work_less_physical TEXT,
        work_careful_physical TEXT,
        work_less_emotional TEXT,
        work_careless_emotional TEXT,

        daily_activity_limit TEXT,

        feel_calm TEXT,
        feel_energetic TEXT,
        feel_sad TEXT,
        social_limit TEXT
    )
    """)

    conn.commit()
    conn.close()

init_db()
# ---------------- Helpers ---------------- #
def classify_pattern(days_per_week: int) -> str:
    return "persistent" if days_per_week >= 4 else "intermittent"

def calculate_follow_up(prev_follow_up, avg_vas, pattern, used_steroid_before):
    # reset condition
    if avg_vas < 5 and pattern == "intermittent":
        return 0

    # first time worsening
    if avg_vas >= 5 and prev_follow_up == 0:
        return 1

    # follow_up = 1 logic
    if prev_follow_up == 1:
        if used_steroid_before == "yes":
            return 2
        return 1

    # follow_up = 2 stays 2
    if prev_follow_up >= 2:
        return 2

    return prev_follow_up

# ---------------- Medicine Algorithm ---------------- #
def generate_recommendation(pattern, avg_vas, follow_up, used_steroid_answer):
    saline = (
        "ล้างจมูกด้วยน้ำเกลือ (Normal saline irrigation)\n"
        "– วันละ 1–2 ครั้ง\n\n"
    )

    oral_ah = (
        "ยาต้านฮิสตามีนชนิดรับประทาน รุ่นที่ 2\n"
        "– วันละ 1 ครั้ง\n\n"
    )

    leuko = (
        "Leukotriene receptor antagonist (LTRA)\n"
        "– วันละ 1 ครั้ง\n\n"
    )

    incs_standard = (
        "ยาสเตียรอยด์พ่นจมูก\n"
        "–1 spray/nostril วันละ 2 ครั้ง กรณี เป็น Budesonide(Rhinocort) , Triamcinolone (Nasacort), Beclomethasone (Beconase) \n"
        "–2 spray/nostril วันละ 1 ครั้ง กรณีเป็น Fluticasone furoate (Avamys),  Mometasone (Nasonex), Fluticasone proprionate (Flixonase)\n"
    )

    incs_high = (
        "ยาสเตียรอยด์พ่นจมูก (เพิ่มขนาดยา)\n"
        "ใช้ ยาพ่นจมูก แบบ ผสม Fluticasone propionate / Azelastine (Dymista)\n"
        "วิธีการใช้ คือ 1 spray/nostril วันละ 2 ครั้ง โดยจะใช้ก็ต่อเมื่อ ถ้าคุณ ใช้ ยาพ่นจมูก steroid 2 กด/ ข้าง วันละ 2 ครั้ง\n"
    )

    # ================= STATE 0 =================
    if follow_up == 0:
        if avg_vas == 0:
            return "อาการของคุณหายดีแล้ว"
        if pattern == "intermittent" and avg_vas < 5:
            return saline + "เลือกอย่างใดอย่างหนึ่ง\n\n" + oral_ah + "หรือ\n\n" + leuko

        if (pattern == "intermittent" and avg_vas >= 5) or \
           (pattern == "persistent" and avg_vas < 5):
            return saline + "เลือกอย่างใดอย่างหนึ่ง\n\n" + oral_ah + "หรือ\n\n" + incs_standard

        if pattern == "persistent" and avg_vas >= 5:
            return saline + incs_standard

    # ================= STATE 1 =================
    if follow_up == 1:
        if avg_vas == 0:
            return "อาการของคุณหายดีแล้ว"
        if avg_vas < 5:
            return "อาการดีขึ้น → ลดระดับยา และใช้ยาต่ออีก 2 สัปดาห์"

        if used_steroid_answer == "no":
            return saline + incs_standard

        return (
            "ส่งพบแพทย์เฉพาะทาง\n"
            "ประเมินการวินิจฉัยและการใช้ยา\n\n"
            + incs_high
        )

    # ================= STATE 2 =================
    if follow_up == 2:
        if avg_vas == 0:
            return "อาการของคุณหายดีแล้ว"
        if avg_vas < 5:
            return "อาการดีขึ้น → ลดระดับยา และใช้ยาต่ออีก 2 สัปดาห์"

        return (
            "ส่งพบแพทย์เฉพาะทาง\n"
            "ประเมินการวินิจฉัยและการใช้ยา\n\n"
            + incs_high
        )

    # ================= STATE 3 =================
    if follow_up == 3:
        if avg_vas < 5:
            return "อาการดีขึ้น → ลดระดับยา และใช้ยาต่ออีก 2 สัปดาห์"
        return (
        "ภูมิคุ้มกันบัมบัดด้วยสารก่อภูมิแพ้\n"
        "ควรได้รับการผ่าตัด"
        )

# ---------------- Routes ---------------- #

@app.route("/", methods=["GET"])
def index():
    return redirect(url_for("welcome"))

@app.route("/welcome")
def welcome():
    return render_template("welcome.html")

## ---------- Update Doctor Notes ---------- #
@app.route("/update_doctor_notes/<int:symptom_id>", methods=["POST"])
def update_doctor_notes(symptom_id):
    if session.get("role") != "doctor":
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    fields = [
        "chlorpheniramine", "other_1st_gen", "cetirizine", "levocetirizine",
        "fexofenadine", "loratadine", "desloratadine", "bilastine", "rupatadine",
        "other_2nd_gen", "pseudoephedrine", "other_oral_decongestant",
        "triprolidine_pseudo", "chlorphen_pseudo", "loratadine_pseudo",
        "montelukast", "immunotherapy_oral", "beclomethasone", "budesonide",
        "fluticasone_propionate", "fluticasone_prop_azelastine", "fluticasone_furoate",
        "mometasone", "triamcinolone", "other_incs", "ephedrine", "oxymetazoline",
        "other_intranasal_decongestant", "azelastine", "levocabastin", "ketotifen",
        "prednisolone", "nasal_irrigation", "other_medications", "immunotherapy_inject",
        "anti_ige", "dupilumab", "benralizumab","olopatadine_mometasone","mepolizumab", 
        "patient_advice", "next_visit"
    ]

    set_clause = ", ".join([f"{f} = %s" for f in fields])
    values = [request.form.get(f, "") for f in fields]
    values.append(datetime.utcnow())
    values.append(symptom_id)

    cur.execute(
        f"UPDATE symptoms SET {set_clause}, doctor_notes_updated_at = %s WHERE id = %s",
        values
    )
    conn.commit()
    conn.close()

    # Get patient_id to redirect back
    patient_id = request.form.get("patient_id")
    flash("Doctor notes saved successfully.", "success")
    return redirect(url_for("patient_detail", patient_id=patient_id))
# ---------- Login ---------- #
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        conn = get_db()
        cur = conn.cursor()

        cur.execute(
            "SELECT * FROM users WHERE username = %s",
            (request.form["username"],)
        )
        user = cur.fetchone()
        conn.close()

        if user and check_password_hash(user["password"], request.form["password"]):
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            return redirect(
                url_for("doctor_dashboard" if user["role"] == "doctor" else "patient_form")
            )

        flash("Invalid login", "danger")

    return render_template("login.html")

# ---------- Signup ---------- #
@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        print("SIGNUP FORM:", dict(request.form))
        conn = get_db()
        cur = conn.cursor()

        try:
            role = request.form["role"]

            # 0️⃣ CHECK DOCTOR CODE
            if role == "doctor":
                if request.form.get("doctor_code") != "SECRET123":
                    flash("Invalid doctor signup code", "danger")
                    return redirect(url_for("signup"))

            # 1️⃣ CREATE USER (RETURNING id is REQUIRED for PostgreSQL)
            cur.execute(
                """
                INSERT INTO users (username, password, role, full_name)
                VALUES (%s, %s, %s, %s)
                RETURNING id
                """,
                (
                    request.form["username"],
                    generate_password_hash(request.form["password"]),
                    role,
                    request.form["full_name"]
                )
            )
            user_id = cur.fetchone()["id"]

            # 2️⃣ PATIENT PROFILE
            if role == "patient":
                cur.execute(
                    """
                    INSERT INTO patient_profiles
                    (user_id, email, phone, address, dob, gender,
                     emergency_contact, insurance_provider, hospital_number)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        user_id,
                        request.form.get("email"),
                        request.form.get("phone"),
                        request.form.get("address"),
                        request.form.get("dob"),
                        request.form.get("gender"),
                        request.form.get("emergency_contact"),
                        request.form.get("insurance_provider"),
                        request.form.get("hospital_number")
                    )
                )

                # 3️⃣ PATIENT HISTORY
                history_data = {
                    "symptom_year_pattern": request.form.get("symptom_year_pattern"),

                    "season_summer": bool(request.form.get("season_summer")),
                    "season_rainy": bool(request.form.get("season_rainy")),
                    "season_winter": bool(request.form.get("season_winter")),
                    "season_summer_rainy": bool(request.form.get("season_summer_rainy")),
                    "season_rainy_winter": bool(request.form.get("season_rainy_winter")),
                    "season_uncertain": bool(request.form.get("season_uncertain")),

                    "duration_per_year": request.form.get("duration_per_year"),
                    "weekly_frequency": request.form.get("weekly_frequency"),

                    "time_6_12": bool(request.form.get("time_6_12")),
                    "time_12_18": bool(request.form.get("time_12_18")),
                    "time_18_24": bool(request.form.get("time_18_24")),
                    "time_24_6": bool(request.form.get("time_24_6")),
                    "time_all_day": bool(request.form.get("time_all_day")),
                    "time_uncertain": bool(request.form.get("time_uncertain")),

                    "living_area": request.form.get("living_area"),
                    "near_road": request.form.get("near_road") == "yes",
                    "housing_type": request.form.get("housing_type"),
                    "air_conditioner": request.form.get("air_conditioner") == "yes",

                    "pet_cat": bool(request.form.get("pet_cat")),
                    "pet_dog": bool(request.form.get("pet_dog")),
                    "pet_bird": bool(request.form.get("pet_bird")),
                    "pet_other": request.form.get("pet_other"),

                    "trigger_dust": bool(request.form.get("trigger_dust")),
                    "trigger_pollen": bool(request.form.get("trigger_pollen")),
                    "trigger_animal": bool(request.form.get("trigger_animal")),
                    "trigger_smoke": bool(request.form.get("trigger_smoke")),
                    "trigger_cold_air": bool(request.form.get("trigger_cold_air")),
                    "trigger_pollution": bool(request.form.get("trigger_pollution")),
                    "trigger_stress": bool(request.form.get("trigger_stress")),
                    "trigger_other": ", ".join(filter(None, [
                        ", ".join(request.form.getlist("trigger_other_list")),
                        request.form.get("trigger_other")
                    ])),

                    "smoking_status": request.form.get("smoking_status"),
                    "cigarettes_per_day": (
                        int(request.form.get("cigarettes_per_day"))
                        if request.form.get("cigarettes_per_day")
                        else None
                    ),
                    "quit_years": (
                        int(request.form.get("quit_years"))
                        if request.form.get("quit_years")
                        else None
                    ),

                    "secondhand_smoke": request.form.get("secondhand_smoke"),

                    "drug_allergy": request.form.get("drug_allergy"),
                    "drug_allergy_name": request.form.get("drug_allergy_name"),
                    "drug_allergy_symptom": request.form.get("drug_allergy_symptom"),

                    "food_allergy": request.form.get("food_allergy"),
                    "food_allergy_name": request.form.get("food_allergy_name"),
                    "food_allergy_symptom": request.form.get("food_allergy_symptom"),

                    "natural_allergy": request.form.get("natural_allergy"),
                    "natural_allergy_symptom": request.form.get("natural_allergy_symptom"),

                    "family_asthma": ",".join(request.form.getlist("family_asthma")),
                    "family_rhinitis": ",".join(request.form.getlist("family_rhinitis")),
                    "family_allergic_conjunctivitis": ",".join(request.form.getlist("family_allergic_conjunctivitis")),
                    "family_atopic_dermatitis": ",".join(request.form.getlist("family_atopic_dermatitis")),

                    "work_performance": request.form.get("work_performance"),
                    "physical_activity_problem": request.form.get("physical_activity_problem"),
                    "stairs_problem": request.form.get("stairs_problem"),

                    "work_less_physical": request.form.get("work_less_physical"),
                    "work_careful_physical": request.form.get("work_careful_physical"),

                    "work_less_emotional": request.form.get("work_less_emotional"),
                    "work_careless_emotional": request.form.get("work_careless_emotional"),

                    "daily_activity_limit": request.form.get("daily_activity_limit"),

                    "feel_calm": request.form.get("feel_calm"),
                    "feel_energetic": request.form.get("feel_energetic"),
                    "feel_sad": request.form.get("feel_sad"),
                    "social_limit": request.form.get("social_limit"),
                }

                columns = ", ".join(history_data.keys())
                placeholders = ", ".join(["%s"] * len(history_data))

                cur.execute(
                    f"""
                    INSERT INTO patient_history (user_id, {columns})
                    VALUES (%s, {placeholders})
                    """,
                    (user_id, *history_data.values())
                )

            conn.commit()

            # Send welcome email
            if role == "patient" and request.form.get("email"):
                try:
                    send_welcome_email(request.form.get("email"), request.form["full_name"])
                except Exception as e:
                    print(f"Error sending welcome email: {e}")

            flash(
                "สมัครสมาชิกสำเร็จ กรุณาเข้าสู่ระบบและกรอกแบบประเมินอาการ\n\n"
                "Signup successful. Please log in and complete the assessment form.",
                "success"
            )
            return redirect(url_for("login"))

        except Exception as e:
            conn.rollback()
            flash(f"Signup error: {e}", "danger")

        finally:
            conn.close()

    return render_template("signup.html")

# ---------- Doctor Dashboard ---------- #
@app.route("/doctor_dashboard")
def doctor_dashboard():
    if session.get("role") != "doctor":
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            u.id,
            u.full_name,
            p.phone,
            p.email,
            COUNT(s.id) AS record_count
        FROM users u
        LEFT JOIN patient_profiles p ON u.id = p.user_id
        LEFT JOIN symptoms s ON u.id = s.user_id
        WHERE u.role = 'patient'
        GROUP BY u.id, u.full_name, p.phone, p.email
        ORDER BY u.full_name
    """)

    patients = cur.fetchall()
    conn.close()

    return render_template("doctor_dashboard.html", patients=patients)


## ---------- Doctor Stats ---------- #
@app.route("/doctor_stats")
def doctor_stats():
    if session.get("role") != "doctor":
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    # total patients
    cur.execute(
        "SELECT COUNT(*) AS c FROM users WHERE role = 'patient'"
    )
    total_patients = cur.fetchone()["c"]

    # gender breakdown
    cur.execute("""
        SELECT COALESCE(p.gender, 'unknown') AS gender, COUNT(*) AS c
        FROM users u
        LEFT JOIN patient_profiles p ON u.id = p.user_id
        WHERE u.role = 'patient'
        GROUP BY COALESCE(p.gender, 'unknown')
    """)
    gender_rows = cur.fetchall()
    genders = {r["gender"] or "unknown": r["c"] for r in gender_rows}

    # latest symptom row per patient
    cur.execute("""
        SELECT s.*
        FROM symptoms s
        WHERE s.id IN (
            SELECT MAX(id)
            FROM symptoms
            GROUP BY user_id
        )
    """)
    latest_rows = cur.fetchall()

    # medicine effect stats (all rows)
    cur.execute("""
        SELECT medicine_effect, COUNT(*) as c
        FROM symptoms
        WHERE medicine_effect IS NOT NULL
        GROUP BY medicine_effect
    """)
    me_rows = cur.fetchall()

    conn.close()

    # compute combos and treatment / VAS counts in Python
    im_mild = im_mod = per_mild = per_mod = 0
    treatments = {
        "oral_antihistamine": 0,
        "incs": 0,
        "ltra": 0,
        "saline": 0,
        "referral": 0
    }
    vas_counts = [0] * 11  # 0..10

    for r in latest_rows:
        pattern = (r["pattern"] or "").lower()
        avg_vas = float(r["avg_vas"]) if r["avg_vas"] is not None else 0.0
        severity = "mild" if avg_vas < 5 else "modsev"

        if pattern == "intermittent":
            if severity == "mild":
                im_mild += 1
            else:
                im_mod += 1
        elif pattern == "persistent":
            if severity == "mild":
                per_mild += 1
            else:
                per_mod += 1

        rec = (r["recommendation"] or "").lower()
        if "ฮิสตามีน" in rec or "antihistamine" in rec or "oral_ah" in rec:
            treatments["oral_antihistamine"] += 1
        if "สเตียรอยด์" in rec or "steroid" in rec or "incs" in rec:
            treatments["incs"] += 1
        if "leukotriene" in rec or "ltra" in rec or "leuko" in rec:
            treatments["ltra"] += 1
        if "ล้างจมูกด้วยน้ำเกลือ" in rec or "normal saline" in rec or "saline" in rec:
            treatments["saline"] += 1
        if "ส่งพบแพทย์" in rec or "refer" in rec or "ผ่าตัด" in rec:
            treatments["referral"] += 1

        v = int(round(avg_vas))
        if v < 0:
            v = 0
        if v > 10:
            v = 10
        vas_counts[v] += 1

    # Prepare medicine effect data [-3 to +3]
    me_map = {i: 0 for i in range(-3, 4)}
    for r in me_rows:
        val = r["medicine_effect"]
        if val in me_map:
            me_map[val] = r["c"]
    medicine_effect_data = [me_map[i] for i in range(-3, 4)]

    combo_counts = [im_mild, im_mod, per_mild, per_mod]

    return render_template(
        "doctor_stats.html",
        total_patients=total_patients,
        genders=genders,
        combo_counts=combo_counts,
        treatments=treatments,
        vas_counts=vas_counts,
        medicine_effect_data=medicine_effect_data
    )

# ---------- Patient Detail ---------- #
@app.route("/patient/<int:patient_id>")
def patient_detail(patient_id):
    if session.get("role") != "doctor":
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    # patient + profile + history
    cur.execute("""
        SELECT 
            u.id,
            u.full_name,

            -- profile
            p.email,
            p.phone,
            p.address,
            p.dob,
            p.gender,
            p.emergency_contact,
            p.insurance_provider,
            p.hospital_number,

            -- history
            h.symptom_year_pattern,
            h.season_summer, h.season_rainy, h.season_winter,
            h.season_summer_rainy, h.season_rainy_winter, h.season_uncertain,
            h.duration_per_year, h.weekly_frequency,
            h.time_6_12, h.time_12_18, h.time_18_24, h.time_24_6, h.time_all_day, h.time_uncertain,
            h.living_area, h.near_road, h.housing_type, h.air_conditioner,
            h.pet_cat, h.pet_dog, h.pet_bird, h.pet_other,
            h.trigger_dust, h.trigger_pollen, h.trigger_animal,
            h.trigger_smoke, h.trigger_cold_air, h.trigger_pollution, h.trigger_stress, h.trigger_other,
            h.smoking_status, h.cigarettes_per_day, h.quit_years, h.secondhand_smoke,
            h.drug_allergy, h.drug_allergy_name, h.drug_allergy_symptom,
            h.food_allergy, h.food_allergy_name, h.food_allergy_symptom,
            h.natural_allergy, h.natural_allergy_symptom,
            h.family_asthma, h.family_rhinitis, h.family_allergic_conjunctivitis, h.family_atopic_dermatitis,
            h.work_performance, h.physical_activity_problem, h.stairs_problem,
            h.work_less_physical, h.work_careful_physical,
            h.work_less_emotional, h.work_careless_emotional,
            h.daily_activity_limit,
            h.feel_calm, h.feel_energetic, h.feel_sad, h.social_limit
        FROM users u
        LEFT JOIN patient_profiles p ON u.id = p.user_id
        LEFT JOIN patient_history h ON u.id = h.user_id
        WHERE u.id = %s
    """, (patient_id,))
    patient = cur.fetchone()

    # all symptom rows
    cur.execute("""
        SELECT * FROM symptoms
        WHERE user_id = %s
        ORDER BY created_at DESC
    """, (patient_id,))
    rows = cur.fetchall()

    # VAS rows (date-wise)
    cur.execute("""
        SELECT DATE(created_at) AS date, avg_vas, recommendation
        FROM symptoms
        WHERE user_id = %s
        ORDER BY DATE(created_at)
    """, (patient_id,))
    vas_rows = cur.fetchall()

    conn.close()

    reports = [{
        "created_at": r["created_at"],
        "tnss": r["tnss"],
        "pattern": r["pattern"],
        "avg_vas": r["avg_vas"],
        "follow_up": r["follow_up"],
        "recommendation": r["recommendation"],
        "data": r["raw_form"] if r["raw_form"] else {},
        "id": r["id"],
        # doctor notes
        "chlorpheniramine": r.get("chlorpheniramine"),
        "other_1st_gen": r.get("other_1st_gen"),
        "cetirizine": r.get("cetirizine"),
        "levocetirizine": r.get("levocetirizine"),
        "fexofenadine": r.get("fexofenadine"),
        "loratadine": r.get("loratadine"),
        "desloratadine": r.get("desloratadine"),
        "bilastine": r.get("bilastine"),
        "rupatadine": r.get("rupatadine"),
        "other_2nd_gen": r.get("other_2nd_gen"),
        "pseudoephedrine": r.get("pseudoephedrine"),
        "other_oral_decongestant": r.get("other_oral_decongestant"),
        "triprolidine_pseudo": r.get("triprolidine_pseudo"),
        "chlorphen_pseudo": r.get("chlorphen_pseudo"),
        "loratadine_pseudo": r.get("loratadine_pseudo"),
        "montelukast": r.get("montelukast"),
        "immunotherapy_oral": r.get("immunotherapy_oral"),
        "beclomethasone": r.get("beclomethasone"),
        "budesonide": r.get("budesonide"),
        "fluticasone_propionate": r.get("fluticasone_propionate"),
        "fluticasone_prop_azelastine": r.get("fluticasone_prop_azelastine"),
        "fluticasone_furoate": r.get("fluticasone_furoate"),
        "mometasone": r.get("mometasone"),
        "triamcinolone": r.get("triamcinolone"),
        "other_incs": r.get("other_incs"),
        "ephedrine": r.get("ephedrine"),
        "oxymetazoline": r.get("oxymetazoline"),
        "other_intranasal_decongestant": r.get("other_intranasal_decongestant"),
        "azelastine": r.get("azelastine"),
        "levocabastin": r.get("levocabastin"),
        "ketotifen": r.get("ketotifen"),
        "prednisolone": r.get("prednisolone"),
        "nasal_irrigation": r.get("nasal_irrigation"),
        "other_medications": r.get("other_medications"),
        "immunotherapy_inject": r.get("immunotherapy_inject"),
        "anti_ige": r.get("anti_ige"),
        "dupilumab": r.get("dupilumab"),
        "benralizumab": r.get("benralizumab"),
        "olopatadine_mometasone": r.get("olopatadine_mometasone"),
"mep    olizumab": r.get("mepolizumab"),
        "patient_advice": r.get("patient_advice"),
        "next_visit": r.get("next_visit"),
        "doctor_notes_updated_at": r.get("doctor_notes_updated_at"),
    } for r in rows]

    return render_template(
        "patient_detail.html",
        patient=patient,
        reports=reports,
        vas_rows=vas_rows
    )

# ---------- Patient Form ---------- #
@app.route("/patient_form", methods=["GET", "POST"])
def patient_form():
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    # latest record
    cur.execute("""
        SELECT * FROM symptoms
        WHERE user_id = %s
        ORDER BY created_at DESC
        LIMIT 1
    """, (session["user_id"],))
    last = cur.fetchone()

    follow_up = last["follow_up"] if last else 0
    need_followup = follow_up in (1, 2)
    next_allowed = None

    if last:
        last_date = last["created_at"]  # already datetime
        next_allowed = last_date + timedelta(days=14)

    # ---------- POST ----------
    if request.method == "POST":
        report_date = datetime.fromisoformat(request.form["report_date"])

        if last and report_date < next_allowed:
            flash(f"กรอกได้อีกครั้งวันที่ {next_allowed:%Y-%m-%d}", "warning")
            return redirect(url_for("patient_form"))

        freq = int(request.form["symptom_frequency"])
        avg_vas = float(request.form["vas_score1"])
        pattern = classify_pattern(freq)
        used_steroid = request.form.get("used_steroid_before", "no")
        prev_follow_up = follow_up

        tnss = (
            int(request.form.get("Frequently sneeze", 0)) +
            int(request.form.get("Stuffed nose", 0)) +
            int(request.form.get("runny nose", 0)) +
            int(request.form.get("itchy nose", 0))
        )

        # 1️⃣ recommendation first
        recommendation = generate_recommendation(
            pattern, avg_vas, prev_follow_up, used_steroid
        )

        # 2️⃣ follow-up logic
        next_follow_up = prev_follow_up
        if avg_vas < 5 and pattern == "intermittent":
            next_follow_up = 0
        elif prev_follow_up == 0 and avg_vas >= 5:
            next_follow_up = 1
        elif prev_follow_up == 1 and avg_vas >= 5:
            next_follow_up = 2 if used_steroid == "yes" else 1
        elif prev_follow_up == 2 and avg_vas >= 5:
            next_follow_up = 3

        # Create dictionary from form data and explicitly add VAS scores
        form_data = {k: request.form.get(k) for k in request.form}
        form_data['vas_score1'] = request.form.get('vas_score1')
        form_data['vas_score2'] = request.form.get('vas_score2')
        form_data['vas_score3'] = request.form.get('vas_score3')

        # Capture multi-select fields from the medication section
        form_data['antihistamine_type'] = request.form.getlist('antihistamine_type')
        form_data['incs_type'] = request.form.getlist('incs_type')

        raw_form = json.dumps(form_data)


        # ----- medicine_effect: update previous row -----
        medicine_effect_answer = request.form.get("medicine_effect")
        if last and medicine_effect_answer:
            try:
                me_val = int(medicine_effect_answer)
                cur.execute(
                    "UPDATE symptoms SET medicine_effect = %s WHERE id = %s",
                    (me_val, last["id"])
                )
            except ValueError:
                pass

        # insert new record
        cur.execute("""
                        INSERT INTO symptoms
                        (user_id, avg_vas, tnss, pattern, recommendation,
                        follow_up, created_at, raw_form, medicine_effect)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        session["user_id"],
                        avg_vas,
                        tnss,
                        pattern,
                        recommendation,
                        next_follow_up,
                        report_date.isoformat(),  # patient date stays
                        raw_form,
                        None
                    ))
        conn.commit()
        conn.close()



        flash("บันทึกข้อมูลเรียบร้อย ดูผลการประเมินที่หน้า Result", "success")
        return redirect(url_for("patient_form", show_result="1"))

    # ================= GET =================
    cur.execute(
        "SELECT * FROM symptoms WHERE user_id = %s ORDER BY created_at DESC",
        (session["user_id"],)
    )
    reports = cur.fetchall()

    cur.execute("""
        SELECT 
            u.full_name,
            p.email,
            p.phone,
            p.gender,
            p.dob,
            p.address
        FROM users u
        LEFT JOIN patient_profiles p ON u.id = p.user_id
        WHERE u.id = %s
    """, (session["user_id"],))
    patient = cur.fetchone()

    conn.close()

    show_medicine_effect_question = bool(last)

    latest_html = ""
    if reports:
        r = reports[0]
        latest_html = Markup(
            f"<b>Date:</b> {r['created_at'].date()}<br>"
            f"<b>Pattern:</b> {r['pattern']}<br>"
            f"<b>VAS:</b> {r['avg_vas']}<br>"
            f"<b>Follow-up:</b> {r['follow_up']}<br>"
            f"<pre>{r['recommendation']}</pre>"
        )

    return render_template(
        "patient_form.html",
        patient=patient,
        reports=reports,
        latest_html=latest_html,
        today=datetime.utcnow().strftime("%Y-%m-%d"),
        need_followup=need_followup,
        show_medicine_effect_question=show_medicine_effect_question
    )

# ---------- Logout ---------- #
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))



if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
